from datetime import datetime
import re
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

def details():
    # Function to get and validate student name
    try:
        s_name = input("Enter the student name:")
        while not re.match(r'^[a-zA-Z\s]+$', s_name):
            if not s_name.isalnum():
                print("You have entered some special characters. Please provide name in alphabets.")
            else:
                print("You have entered numbers. Please provide name in alphabets.")
            s_name = input("Enter the student name:")
    except Exception as e:
        print(f"An error occurred: {e}")

    # Student register no
    try:
        s_regno = input("\nEnter your register number:")   
        while not s_regno.isdigit() or len(s_regno) != 6:
            if s_regno.isalpha():
                print("You have entered alphabets. Please enter only digits.")
            elif s_regno.isalnum():
                print("You have entered a combination of letters and digits. Please enter only digits.")
            else:
                print("You have entered special characters. Please enter only digits.")
            s_regno = input("\nEnter your register number:")
    except Exception as e:
        print(f"An error occurred: {e}")

    # Student date of birth
    try:
        while True:
            stud_dob = input("Enter the date of birth (DD-MM-YYYY):")
            format_of_dob = r'^\d{2}-\d{2}-\d{4}$'
            if re.match(format_of_dob, stud_dob):
                break
            else:
                print("Invalid format of date of birth.")
    except Exception as e:
        print(f"An error occurred: {e}")

    gender = input("Enter the gender (Male/Female):")

    stud_marks=[]
    for i in range(5):
        if i==0:
            #language 1 mark
            while True: 
                try:
                    lang1_mark=int(input("Enter the language 1 mark:"))
                    if lang1_mark<0 or lang1_mark>100:
                        raise ValueError("Marks should be within 0-100")
                    else:
                        stud_marks.append(lang1_mark)
                        break
                except ValueError as ve:
                    print(ve)
        elif i==1:
            #language 2 mark
            while True: 
                try:
                    lang2_mark=int(input("Enter the language 2 mark:"))
                    if lang2_mark<0 or lang2_mark>100:
                        raise ValueError("Marks should be within 0-100")
                    else:
                        stud_marks.append(lang2_mark)
                        break
                except ValueError as ve:
                    print(ve) 
        elif i==2:
            #maths mark        
            while True: 
                try:
                    maths_mark=int(input("Enter the maths mark:"))
                    if maths_mark<0 or maths_mark>100:
                        raise ValueError("Marks should be within 0-100")
                    else:
                        stud_marks.append(maths_mark)
                        break
                except ValueError as ve:
                    print(ve)
        elif i==3:
            #science mark
            while True: 
                try:
                    sci_mark=int(input("Enter the science mark:"))
                    if sci_mark<0 or sci_mark>100:
                        raise ValueError("Marks should be within 0-100")
                    else:
                        stud_marks.append(sci_mark)
                        break
                except ValueError as ve:
                    print(ve) 
        elif i==4:
            #social science mark
            while True: 
                try:
                    socsci_mark=int(input("Enter the social science mark:"))
                    if socsci_mark<0 or socsci_mark>100:
                        raise ValueError("Marks should be within 0-100")
                    else:
                        stud_marks.append(socsci_mark)
                        break
                except ValueError as ve:
                    print(ve)         
    #print(stud_marks)  
    
    grades=[]
    for i in stud_marks:
        if 90<i<=100:
            grades.append("O")
        elif 80<i<=90:
            grades.append("A+")
        elif 70<i<=80:
            grades.append("A")
        elif 60<i<=70:
            grades.append("B+")
        elif 50<i<=60:
            grades.append("C")
        elif 35<i<=50:
            grades.append("D")
        else:
            grades.append("E")         
            
    #print(grades)       
    
    stud_total=0
    stud_total=lang1_mark+lang2_mark+maths_mark+sci_mark+socsci_mark
    
    stud_percent=(stud_total/500)*100
    #pass or fail
    
    try:
        if lang1_mark<35 or lang2_mark<35 or maths_mark<35 or sci_mark<35 or socsci_mark<35:
            final_result="Fail"
        else:
            final_result="Pass"
    except ValueError as ve:
        print(ve)
    save_to_excel(s_name, s_regno, stud_dob, gender, stud_marks,grades, stud_total, stud_percent, final_result)

def save_to_excel(s_name, s_regno, stud_dob, gender, stud_marks, grades, stud_total, stud_percent, final_result):
    file_path = "student_records_in_sheets.xlsx"
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()

    if s_regno in wb.sheetnames:
        sheet = wb[s_regno]
    else:
        sheet = wb.create_sheet(title=s_regno)
        headers = [
            "Student Name", "Register Number", "Date of Birth", "Gender",
            "Language 1", "Language 1 grade", "Language 2", "Language 2 grade",
            "Maths", "Maths grade", "Science", "Science grade", "Social Science",
            "Social Science grade", "Total", "Percentage", "Final Result"
        ]
        values = [
            s_name, s_regno, stud_dob, gender, stud_marks[0], grades[0],
            stud_marks[1], grades[1], stud_marks[2], grades[2], stud_marks[3],
            grades[3], stud_marks[4], grades[4], stud_total, stud_percent, final_result
        ]
        for i in range(len(headers)):
            sheet.append([headers[i], values[i]])
    
    wb.save(file_path)
    print(f"Details have been saved to {file_path}")

def get_details():
    def regno():
        return input("Enter the register no (only 6 digits): ")

    stu_regno = regno()
    while not stu_regno.isdigit() or len(stu_regno) != 6:
        print("Invalid input. Please enter a 6-digit register number.")
        stu_regno = regno()

    while True:
        stu_dob = input("Enter the date of birth (DD-MM-YYYY):")
        format_of_dob = r'^\d{2}-\d{2}-\d{4}$'
        if re.match(format_of_dob, stu_dob):
            break
        else:
            print("Invalid format of date of birth. Please use DD-MM-YYYY format.")

    file = "student_records_in_sheets.xlsx"
    student_found = False

    if os.path.exists(file):
        wb = load_workbook(file)
        if stu_regno in wb.sheetnames:
            ws = wb[stu_regno]
            details = {}
            for row in ws.iter_rows(values_only=True):
                details[row[0]] = row[1]
            if details["Register Number"] == stu_regno and details["Date of Birth"] == stu_dob:
                print("SSLC EXAMINATION:")
                print("Student details:")
                for key, value in details.items():
                    print(f"{key}: {value}")
                student_found = True
                p = input("Do you want to print (yes/no)? ")
                if p.lower() == "yes":
                    print_details(details)
        if not student_found:
            print("No student details are found.")
    else:
        print("No student records found.")

def print_details(details):
    filename = "Students folder"
    if not os.path.exists(filename):
        os.makedirs(filename)

    file = f"{details['Register Number']}.txt"
    filepath = os.path.join(filename, file)

    with open(filepath, "w") as f:
        f.write("SSLC EXAMINATION:\n")
        f.write("Student details:\n")
        for key, value in details.items():
            f.write(f"{key}: {value}\n")

    os.startfile(filepath)
# Starting of the program
def start():
    while True:
        y = input("Do you want to enter student details or search for the student details? (enter/search/exit): ")
        if y == "enter":
            details()
        elif y == "search":
            get_details()
        elif y == "exit":
            break
        else:
            print("Invalid input. Please give input as 'enter', 'search', 'exit'.")

start()