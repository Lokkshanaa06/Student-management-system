from datetime import datetime
import re
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font,PatternFill 
from tabulate import tabulate

def details():
    # Function to get and validate student name
    #student name
    try:
        s_name=input("Enter the student name:")
        while not re.match(r'^[a-zA-Z\s]+$', s_name):
                if not s_name.isalnum():
                    print("You have entered some special characters. please provide name in alphabets")
                    s_name=input("Enter the student name:")
                else:
                    if not s_name.replace(' ', '').isalpha():
                        print("You have entered numbers.please provide name in alphabet")
                    s_name=input("Enter the student name:")
    except Exception as e:
        print(f"An error occured {e}")

    #student register no
    try:
        s_regno=input("\nEnter your register number:")   
        while not s_regno.isdigit():
            if s_regno.isalpha():
                print("You have entered alphabets.Please enter only digits")
                s_regno=input("\nEnter your register number:")  
            elif s_regno.isalnum():
                print("You have entered the combination of letters and digits.Please enter only digits")
                s_regno=input("\nEnter your register number:") 
            else:
                print("You have entered special characters.Please enter only digits")
                s_regno=input("\nEnter your register number:") 
        while len(s_regno)!=6:
            print("Reguster number should contain 6 digits only.")
            s_regno=input("\nEnter your register number:") 
    except Exception as e:
        print(f"An error occured {e}")

    #student date of birth
    try:
        while True:
            stud_dob=input("Enter the date of birth:")
            format_of_dob=r'^\d{2}-\d{2}-\d{4}$'
            if re.match(format_of_dob,stud_dob):
                #print("Date of birth:",stud_dob)
                break
            else:
                print("Invalid format of date of birth") 
    except Exception as e:
        print(f"An error occured {e}")

    gender=input("Enter the gender (Male/Female):")
    #print("The gender of the student is:",gender)
    # marks array
    stud_marks=[]
    for i in range(5):
        if i==0:
            #language 1 mark
            while True: 
                try:
                    lang1_mark=int(input("Enter the language 1 mark:"))
                    if lang1_mark<0 or lang1_mark>100:
                        raise ValueError("Marks should be within 0-100")
                    else:
                        stud_marks.append(lang1_mark)
                        break
                except ValueError as ve:
                    print(ve)
        elif i==1:
            #language 2 mark
            while True: 
                try:
                    lang2_mark=int(input("Enter the language 2 mark:"))
                    if lang2_mark<0 or lang2_mark>100:
                        raise ValueError("Marks should be within 0-100")
                    else:
                        stud_marks.append(lang2_mark)
                        break
                except ValueError as ve:
                    print(ve) 
        elif i==2:
            #maths mark        
            while True: 
                try:
                    maths_mark=int(input("Enter the maths mark:"))
                    if maths_mark<0 or maths_mark>100:
                        raise ValueError("Marks should be within 0-100")
                    else:
                        stud_marks.append(maths_mark)
                        break
                except ValueError as ve:
                    print(ve)
        elif i==3:
            #science mark
            while True: 
                try:
                    sci_mark=int(input("Enter the science mark:"))
                    if sci_mark<0 or sci_mark>100:
                        raise ValueError("Marks should be within 0-100")
                    else:
                        stud_marks.append(sci_mark)
                        break
                except ValueError as ve:
                    print(ve) 
        elif i==4:
            #social science mark
            while True: 
                try:
                    socsci_mark=int(input("Enter the social science mark:"))
                    if socsci_mark<0 or socsci_mark>100:
                        raise ValueError("Marks should be within 0-100")
                    else:
                        stud_marks.append(socsci_mark)
                        break
                except ValueError as ve:
                    print(ve)         
    #print(stud_marks)  
    
    grades=[]
    for i in stud_marks:
        if 90<i<=100:
            grades.append("O")
        elif 80<i<=90:
            grades.append("A+")
        elif 70<i<=80:
            grades.append("A")
        elif 60<i<=70:
            grades.append("B+")
        elif 50<i<=60:
            grades.append("C")
        elif 35<i<=50:
            grades.append("D")
        else:
            grades.append("E")         
            
    #print(grades)       
    
    stud_total=0
    stud_total=lang1_mark+lang2_mark+maths_mark+sci_mark+socsci_mark
    
    stud_percent=(stud_total/500)*100
    #pass or fail
    
    try:
        if lang1_mark<35 or lang2_mark<35 or maths_mark<35 or sci_mark<35 or socsci_mark<35:
            final_result="Fail"
        else:
            final_result="Pass"
    except ValueError as ve:
        print(ve)
    save_to_excel(s_name, s_regno, stud_dob, gender, stud_marks,grades, stud_total, stud_percent, final_result)

def save_to_excel(s_name, s_regno, stud_dob, gender, stud_marks, grades, stud_total, stud_percent, final_result):
    
    
    file_path = "student_records_with_grades.xlsx"
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Student Details"
        sheet.append(["Student Name", "Register Number", "Date of Birth", "Gender", "Language 1", "Language 1 grade", "Language 2", "Language 2 grade", "Maths", "Maths grade" ,"Science" , "Science grade" , "Social Science" , "Social Science grade" , "Total", "Percentage", "Final Result"])
        #bold the column headexit
        #for cell in sheet["1:1"]:
            #cell.font=Font(bold=True)
    for cell in sheet["1:1"]:
        cell.font=Font(bold=True)
        
    
    sheet.append([s_name, s_regno, stud_dob, gender, stud_marks[0], grades[0], stud_marks[1], grades[1], stud_marks[2], grades[2], stud_marks[3], grades[3], stud_marks[4], grades[4], stud_total, stud_percent, final_result])
    for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,min_col=17,max_col=17):
        for cell in row:
            if cell.value=="Fail":
                cell.fill=PatternFill(start_color="FFFF00",end_color="FFFF00",fill_type="solid")
    
    wb.save(file_path)
    print(f"Details have been saved to {file_path}")

def get_details():
    
    try:
        s_regno=input("\nEnter your register number:")   
        while not s_regno.isdigit():
            if s_regno.isalpha():
                print("You have entered alphabets.Please enter only digits")
                s_regno=input("\nEnter your register number:")  
            elif s_regno.isalnum():
                print("You have entered the combination of letters and digits.Please enter only digits")
                s_regno=input("\nEnter your register number:") 
            else:
                print("You have entered special characters.Please enter only digits")
                s_regno=input("\nEnter your register number:") 
        while len(s_regno)!=6:
            print("Reguster number should contain 6 digits only.")
            s_regno=input("\nEnter your register number:") 
    except Exception as e:
        print(f"An error occured {e}")

    #student date of birth
    try:
        while True:
            stud_dob=input("Enter the date of birth:")
            format_of_dob=r'^\d{2}-\d{2}-\d{4}$'
            if re.match(format_of_dob,stud_dob):
                #print("Date of birth:",stud_dob)
                break
            else:
                print("Invalid format of date of birth") 
    except Exception as e:
        print(f"An error occured {e}")

    file = "student_records_with_grades.xlsx"
    student_found = False

    if os.path.exists(file):
        wb = load_workbook(file)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            if row[1] == s_regno and row[2] == stud_dob :
                print("SSLC EXAMINATION:")
                print("Student details:")
                print("Student Name:    ", row[0])
                print("Register Number: ", row[1])
                print("Date of Birth:   ", row[2])
                print("Gender:          ", row[3])
                table_details=[
                    ["Language 1",row[4],row[5]],
                    ["Language 2",row[6],row[7]],
                    ["Maths",row[8],row[9]],
                    ["Science",row[10],row[11]],
                    ["Social Science",row[12],row[13]],
                    ["Total",row[14],""],
                    ["Percentage",row[15],""],
                    ["Result",row[16],""],   
                ]
                print("\n Marks and grades")
                print(tabulate(table_details,headers=["Subjects","Marks","Grades"],tablefmt="grid"))
                
                student_found = True
                p=input("Do you want to print(yes/no)? ")
                if p=="yes":
                    print_details(row)
                break
            #print("The student details are not found")
    if not student_found:
        print("No student details are found.")
    

def print_details(row):
    filename="Students folder"
    if not os.path.exists(filename):
        os.makedirs(filename)
    
    file=f"{row[1]}.txt"
    filepath=os.path.join(filename,file)
    
    with open (filepath,"w") as f:
                f.write("SSLC EXAMINATION:")
                f.write("\nStudent details:")
                f.write(f"\nStudent Name:     {row[0]}")
                f.write(f"\nRegister Number:  {row[1]}")
                f.write(f"\nDate of Birth:    {row[2]}")
                f.write(f"\nGender:           {row[3]}")
                table_details=[
                    ["Language 1",row[4],row[5]],
                    ["Language 2",row[6],row[7]],
                    ["Maths",row[8],row[9]],
                    ["Science",row[10],row[11]],
                    ["Social Science",row[12],row[13]],
                    ["Total",row[14],""],
                    ["Percentage",row[15],""],
                    ["Result",row[16],""],   
                ]
                f.write("\n Marks and grades")
                f.write(tabulate(table_details,headers=["Subjects","Marks","Grades"],tablefmt="grid"))
    
    os.startfile(filepath)

# starting of the program
def start():
    while True:
        y = input("Do you want to enter student details or search for the student details? (enter/search/exit): ")
        if y == "enter":
            details()
        elif y=="search":
            get_details()
        elif y == "exit":
            break
        else:
            print("Invalid input. Please give input as 'enter', 'search', 'exit'")
start()
