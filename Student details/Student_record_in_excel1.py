from datetime import datetime
import re
import os
from openpyxl import Workbook, load_workbook

def get_details():
    
    def regno():
        stud_regno = input("Enter the register no (only 6 digits): ")
        return stud_regno
    stu_regno = input("Enter the student register no:")
    while not stu_regno.isdigit() or len(stu_regno) != 6:
        if stu_regno.isalpha():
            print("You have entered alphabets. Enter digits.")
        elif stu_regno.isalnum():
            print("You have entered a combination of alphabets and numbers. Enter digits.")
        else:
            print("You have entered special characters. Enter digits.")
        stu_regno = regno()
    
    while True:
        stu_dob = input("Enter the date of birth (DD-MM-YYYY):")
        format_of_dob = r'^\d{2}-\d{2}-\d{4}$'
        if re.match(format_of_dob, stu_dob):
            break
        else:
            print("Invalid format of date of birth. Please use DD-MM-YYYY format.")
    
    file = "student_records.xlsx"
    student_found = False

    if os.path.exists(file):
        wb = load_workbook(file)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            if row[1] == stu_regno and row[2] == stu_dob :
                print("SSLC EXAMINATION:")
                print("Student details:")
                print("Student Name:    ", row[0])
                print("Register Number: ", row[1])
                print("Date of Birth:   ", row[2])
                print("Gender:          ", row[3])
                print("Language 1:      ", row[4])
                print("Language 2:      ", row[5])
                print("Maths:           ", row[6])
                print("Science:         ", row[7])
                print("Social Science:  ", row[8])
                print("Total:           ", row[9])
                print("Percentage:      ", row[10])
                print("Final Result:    ", row[11])
                student_found = True
                p=input("Do you want to print(yes/no)? ")
                if p=="yes":
                    print_details(row)
                break
            #print("The student details are not found")
    if not student_found:
        print("No student details are found.")
    

def print_details(row):
    filename="Students folder"
    if not os.path.exists(filename):
        os.makedirs(filename)
    
    file=f"{row[1]}.txt"
    filepath=os.path.join(filename,file)
    
    with open (filepath,"w") as f:
        f.write("SSLC EXAMINATION:\n")
        f.write("Student details:\n")
        f.write(f"Student Name:    {row[0]}\n")
        f.write(f"Register Number: {row[1]}\n")
        f.write(f"Date of Birth:   {row[2]}\n")
        f.write(f"Gender:          {row[3]}\n")
        f.write(f"Language 1:      {row[4]}\n")
        f.write(f"Language 2:      {row[5]}\n")
        f.write(f"Maths:           {row[6]}\n")
        f.write(f"Science:         {row[7]}\n")
        f.write(f"Social Science:  {row[8]}\n")
        f.write(f"Total:           {row[9]}\n")
        f.write(f"Percentage:      {row[10]}\n")
        f.write(f"Final Result:    {row[11]}\n")
    
    os.startfile(filepath)
    

while True:
    s=input("Do you want to search for the student details(search/exit)?")
    if s=="search":
        get_details()
    elif s=="exit":
        break
    else:
        print("Invalid input")   

