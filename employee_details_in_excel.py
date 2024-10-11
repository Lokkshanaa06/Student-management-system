from  openpyxl import Workbook , load_workbook
import os
def get_details():
    emp_name=input("Enter the employee name:")
    emp_id=input("Enter the employee id:")
    emp_designation=input("Enter the designation of the employee:")
    emp_salary=int(input("Enter the salary of the employee:"))
    emp_gender=input("Enter the gender:")
    insert_record(emp_name,emp_id,emp_designation,emp_salary,emp_gender)
def insert_record(e_name,e_id,e_desig,e_salary,e_gender):
    file="employee_sheet.xlsx"
    if os.path.exists(file):
        wb=load_workbook(file)
        sheet=wb.active
    else:
        wb= Workbook()
        sheet=wb.active
        sheet.title ="employee details"
        sheet.append(["employee name","employee id","employee designation","employee salary","employee gender"])
    sheet.append([e_name,e_id,e_desig,e_salary,e_gender])
    wb.save(file)
    print("Details have ben saved")
while True:
    s=input("Do you want enter details (yes/no)?")
    if s=="yes":
        get_details()
    else:
        break