import os
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill

file="student_records.xlsx"

if os.path.exists(file):
    wb=load_workbook(file)
    sheet=wb.active
    
    for cell in sheet["1:1"]:
        cell.font=Font(bold=True)
        
    for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,min_col=12,max_col=12):
        for cell in row:
            if cell.value=="Fail":
                cell.fill=PatternFill(start_color="FFFF00",end_color="FFFF00",fill_type="solid")
    wb.save(file)
    print("Required Changes have been made")
    '''
    r=sheet.max_row
    c=sheet.max_column
    #prints the value of the first column
    print("values of first column")
    for i in range(1,r+1):
        v=sheet.cell(row=i,column=1)
        print(v.value)
        
    #prints the value of the first row
    print("values of first row")
    for j in range(1,c+1):
        v=sheet.cell(row=1,column=j)
        print(v.value)
    
    #prints the multiple values of the cells
    cell=sheet['A1':'B6']
    for cell1,cell2 in cell:
        print(cell1.value,cell2.value)
    '''
else:
    print("Changes have not been done")