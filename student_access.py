import re
from openpyxl import workbook,load_workbook
import os

def get_details(stud_regno,stud_dob):
    #filename of the excel sheet
    file="student_records.xlsx"
    student_found=False
    
    if os.path.exists(file):
        wb=load_workbook(file)
        sheet=wb.active
        for row in sheet.iter_rows(values_only=True):
            if row[1]==stud_regno and row[2]==stud_dob:
                print("SSLC EXAMINATION RESULTS")
                print("------------------------------------------------------")
                print(f"Name:               {row[0]}")
                print(f"Register number:    {row[1]}")
                print(f"Date of birth:      {row[2]}")
                print(f"Gender:             {row[3]}")
                print(f"Language 1:         {row[4]}")
                print(f"Language 2:         {row[5]}")
                print(f"Maths:              {row[6]}")
                print(f"Science:            {row[7]}")
                print(f"Social Science:     {row[8]}")
                print(f"Total:              {row[9]}")
                print(f"Percentage:         {row[10]}")
                print(f"Final Result:       {row[11]}")
                student_found=True
    if not student_found:
        print("Student details are not found")
    

def start():
    
    #input student register number
    try:
        s_regno=input("\nEnter your register number:")   
        while not s_regno.isdigit():
            if s_regno.isalpha():
                print("You have entered alphabets.Please enter only digits")
                s_regno=input("\nEnter your register number:")  
            elif s_regno.isalnum():
                print("You have entered the combination of letters and digits.Please enter only digits")
                s_regno=input("\nEnter your register number:") 
            else:
                print("You have entered special characters.Please enter only digits")
                s_regno=input("\nEnter your register number:") 
        while len(s_regno)!=6:
            print("Reguster number should contain 6 digits only.")
            s_regno=input("\nEnter your register number:") 
    except Exception as e:
        print(f"An error occured {e}")
  
    #input student date of birth
    try:
        while True:
            s_dob=input("\nEnter your date of birth:")
            format=r'^\d{2}-\d{2}-\d{4}$'
            if re.match(format,s_dob):
                break
            else:
                print("Invaild format of the date of birth")
    except Exception as e:
        print(f"An error occured {e}")
    #get_details(s_regno,s_dob)
    
    get_details(s_regno,s_dob)

try:
    while True:    
            s=input("Do you want to enter your details (yes/no)\n")
            if s.lower()=="yes":
                start()
            elif s.lower()=="no":
                break
            else:
                print("Invalid input")
except Exception as e:
        print(f"An error occured {e}")

